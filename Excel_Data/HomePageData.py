import openpyxl
import os

ROOT_DIR = os.path.abspath(os.curdir)


class HomePageData:

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(ROOT_DIR + "\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # for rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # for columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]

