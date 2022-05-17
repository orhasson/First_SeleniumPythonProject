import openpyxl
import os

ROOT_DIR = os.path.abspath(os.curdir)

## This is a demo of the working with Excel

book = openpyxl.load_workbook(ROOT_DIR + "\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Or"

print(sheet.cell(row=2, column=2).value)
print(sheet.max_column)
print(sheet.max_row)
print(sheet['A3'].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase1":

        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
